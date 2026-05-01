import csv
import json as _json
import re
import urllib.parse
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Case, DecimalField, ExpressionWrapper, F, Sum, Value, When
from django.db.models.functions import Coalesce
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_GET, require_POST

from clients.models import Client
from products.models import Product
from services.models import Service

from .models import Payment, Sale, SaleItem, SaleModificationLog


# ── Helpers ──────────────────────────────────────────────────────────────────

def _parse_decimal(value, default=Decimal("0.00")):
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return default


# ── sale_list ─────────────────────────────────────────────────────────────────

@login_required
def sale_list(request):
    periode = request.GET.get("periode", "today")
    today = date.today()
    start = end = None

    if periode == "week":
        start = today - timedelta(days=today.weekday())
        end = today
    elif periode == "month":
        start = today.replace(day=1)
        end = today
    elif periode == "all":
        pass  # pas de filtre date
    else:
        periode = "today"
        start = today
        end = today

    base_qs = Sale.objects.select_related("client", "created_by").prefetch_related("items")
    if start is not None:
        base_qs = base_qs.filter(sale_date__gte=start, sale_date__lte=end)

    # Pagination (25 par page)
    paginator = Paginator(base_qs, 25)
    page = paginator.get_page(request.GET.get("page", 1))
    sales = page  # objet Page → itérable + métadonnées de pagination

    # Agrégats financiers : uniquement ventes actives sur la période
    active_filter = {"status": Sale.SaleStatus.ACTIVE}
    if start is not None:
        active_filter["sale_date__gte"] = start
        active_filter["sale_date__lte"] = end
    active_sales = Sale.objects.filter(**active_filter)

    aggregates = active_sales.aggregate(
        total_ca=Sum("total_amount"),
        total_profit=Sum("total_profit"),
    )
    total_ca = aggregates["total_ca"] or Decimal("0.00")
    total_profit = aggregates["total_profit"] or Decimal("0.00")

    # Répartition produits / prestations (ventes actives seulement)
    items_filter = {"sale__status": Sale.SaleStatus.ACTIVE}
    if start is not None:
        items_filter["sale__sale_date__gte"] = start
        items_filter["sale__sale_date__lte"] = end
    items_qs = SaleItem.objects.filter(**items_filter)

    prod_agg = items_qs.filter(product__isnull=False).aggregate(
        ca=Sum("line_total"), profit=Sum("line_profit")
    )
    svc_agg = items_qs.filter(service__isnull=False).aggregate(
        ca=Sum("line_total"), profit=Sum("line_profit")
    )
    rev_products = prod_agg["ca"] or Decimal("0.00")
    profit_products = prod_agg["profit"] or Decimal("0.00")
    rev_services = svc_agg["ca"] or Decimal("0.00")
    profit_services = svc_agg["profit"] or Decimal("0.00")

    # Répartition par catégorie
    from products.models import ProductCategory
    cat_breakdown = []
    for cat in ProductCategory.objects.order_by("name"):
        agg = items_qs.filter(product__category=cat).aggregate(
            ca=Sum("line_total"), profit=Sum("line_profit")
        )
        cat_breakdown.append({
            "slug": cat.slug,
            "name": cat.name,
            "ca": agg["ca"] or Decimal("0.00"),
            "profit": agg["profit"] or Decimal("0.00"),
        })

    # Encaissements réels
    _zero = Value(Decimal("0.00"))
    _df = DecimalField(max_digits=14, decimal_places=2)
    pay_filter = {}
    if start is not None:
        pay_filter["payment_date__gte"] = start
        pay_filter["payment_date__lte"] = end
    payments_qs = Payment.objects.filter(**pay_filter).annotate(
        prop_profit=Case(
            When(
                sale__total_amount__gt=0,
                then=ExpressionWrapper(
                    F("amount") * F("sale__total_profit") / F("sale__total_amount"),
                    output_field=_df,
                ),
            ),
            default=_zero,
            output_field=_df,
        )
    )
    paid_agg = payments_qs.aggregate(
        ca=Coalesce(Sum("amount"), _zero, output_field=_df),
        profit=Coalesce(Sum("prop_profit"), _zero, output_field=_df),
    )
    paid_ca = paid_agg["ca"]
    paid_profit = paid_agg["profit"]

    return render(request, "sales/list.html", {
        "sales": sales,
        "page_obj": page,
        "period": periode,
        "today": today,
        "total_ca": total_ca,
        "total_profit": total_profit,
        "rev_products": rev_products,
        "profit_products": profit_products,
        "rev_services": rev_services,
        "profit_services": profit_services,
        "cat_breakdown": cat_breakdown,
        "paid_ca": paid_ca,
        "paid_profit": paid_profit,
    })


# ── sale_create ───────────────────────────────────────────────────────────────

@login_required
def sale_create(request):
    products = Product.objects.filter(is_active=True).select_related("category").order_by("name")
    services = Service.objects.filter(is_active=True).order_by("name")

    products_data = [
        {
            "id": p.id,
            "name": p.name,
            "category": p.category.name,
            "category_key": p.category.slug,
            "selling_price": str(p.selling_price),
            "wholesale_price": str(p.wholesale_price) if p.wholesale_price is not None else None,
            "purchase_price": str(p.purchase_price),
            "stock_quantity": p.stock_quantity,
            "is_out_of_stock": p.is_out_of_stock,
        }
        for p in products
    ]
    services_data = [
        {
            "id": s.id,
            "name": s.name,
            "price": str(s.price),
        }
        for s in services
    ]

    if request.method == "POST":
        cart_raw = request.POST.get("cart_data", "")
        client_id = request.POST.get("client_id", "").strip()
        payment_amount_raw = request.POST.get("payment_amount", "").strip()
        payment_method = request.POST.get("payment_method", "cash")

        # Parse cart
        try:
            cart = _json.loads(cart_raw) if cart_raw else []
        except _json.JSONDecodeError:
            cart = []

        if not cart:
            return render(request, "sales/create.html", {
                "products_data": products_data,
                "services_data": services_data,
                "error": "Le panier est vide.",
            })

        # Resolve client
        client = None
        if client_id:
            try:
                client = Client.objects.get(pk=int(client_id), is_active=True)
            except (Client.DoesNotExist, ValueError):
                client = None

        # Validate: services require a client
        has_services = any(item.get("type") == "service" for item in cart)
        if has_services and not client:
            return render(request, "sales/create.html", {
                "products_data": products_data,
                "services_data": services_data,
                "error": "Un client est requis pour enregistrer des prestations de service.",
            })

        try:
            with transaction.atomic():
                from services.models import ServiceExecution

                sale = Sale.objects.create(
                    client=client,
                    created_by=request.user,
                    sale_date=date.today(),
                )

                for item in cart:
                    item_type = item.get("type")
                    item_id = item.get("id")
                    qty = int(item.get("qty", 1))
                    unit_price = _parse_decimal(item.get("unit_price"))
                    purchase_price = _parse_decimal(item.get("purchase_price", "0"))

                    if item_type == "product":
                        product = Product.objects.select_for_update().get(pk=item_id)
                        SaleItem.objects.create(
                            sale=sale,
                            product=product,
                            quantity=qty,
                            unit_price=unit_price,
                            purchase_price_snapshot=purchase_price,
                        )
                    elif item_type == "service":
                        service_obj = Service.objects.get(pk=item_id)
                        tours_per_month_raw = item.get("tours_per_month")
                        tours_per_month = int(tours_per_month_raw) if tours_per_month_raw else None
                        sale_item = SaleItem.objects.create(
                            sale=sale,
                            service=service_obj,
                            quantity=qty,
                            unit_price=unit_price,
                            purchase_price_snapshot=Decimal("0.00"),
                        )
                        # Premier passage — lié au SaleItem
                        first_exec = ServiceExecution.objects.create(
                            client=client,
                            service=service_obj,
                            sale_item=sale_item,
                            execution_date=sale.sale_date,
                            tours_per_month=tours_per_month,
                        )
                        # Passages suivants pré-planifiés si le client paye plusieurs tours
                        if qty > 1 and tours_per_month:
                            from datetime import timedelta
                            interval = first_exec.interval_days() or 0
                            prev_date = sale.sale_date
                            for i in range(1, qty):
                                raw_date = prev_date + timedelta(days=interval)
                                # Ajuste au même jour de semaine que le premier passage
                                offset = (sale.sale_date.weekday() - raw_date.weekday()) % 7
                                next_date = raw_date + timedelta(days=offset)
                                prev_date = next_date
                                ServiceExecution.objects.create(
                                    client=client,
                                    service=service_obj,
                                    tours_per_month=tours_per_month,
                                    execution_date=next_date,
                                    parent_execution=first_exec,
                                )

                sale.recompute_totals()

                # Optional immediate payment
                payment_amount = _parse_decimal(payment_amount_raw)
                if payment_amount > Decimal("0.00"):
                    # Cap at total
                    if payment_amount > sale.total_amount:
                        payment_amount = sale.total_amount
                    valid_methods = [m[0] for m in Payment.Method.choices]
                    if payment_method not in valid_methods:
                        payment_method = "cash"
                    Payment.objects.create(
                        sale=sale,
                        recorded_by=request.user,
                        amount=payment_amount,
                        payment_method=payment_method,
                        payment_date=date.today(),
                    )

        except ValueError as e:
            return render(request, "sales/create.html", {
                "products_data": products_data,
                "services_data": services_data,
                "error": str(e),
            })
        except Exception as e:
            return render(request, "sales/create.html", {
                "products_data": products_data,
                "services_data": services_data,
                "error": f"Erreur lors de la création de la vente : {e}",
            })

        return redirect("sales:detail", pk=sale.pk)

    return render(request, "sales/create.html", {
        "products_data": products_data,
        "services_data": services_data,
    })


# ── sale_detail ───────────────────────────────────────────────────────────────

@login_required
def sale_detail(request, pk):
    sale = get_object_or_404(
        Sale.objects.select_related("client", "created_by")
        .prefetch_related("items__product", "items__service", "payments__recorded_by"),
        pk=pk,
    )

    whatsapp_text = _build_whatsapp_message(sale)
    encoded_text = urllib.parse.quote(whatsapp_text)

    whatsapp_url = f"https://api.whatsapp.com/send?text={encoded_text}"
    whatsapp_client_url = None  # URL directe avec le numéro du client
    if sale.client and sale.client.phone:
        phone = re.sub(r"\D", "", sale.client.phone)
        if phone.startswith("00"):
            phone = phone[2:]
        elif phone.startswith("0"):
            phone = "228" + phone[1:]
        if phone:
            whatsapp_url = f"https://wa.me/{phone}?text={encoded_text}"
            whatsapp_client_url = f"https://wa.me/{phone}?text={urllib.parse.quote('Bonjour, veuillez trouver ci-joint votre facture AquaTogo n°' + str(sale.pk).zfill(4) + '.')}"

    return render(request, "sales/detail.html", {
        "sale": sale,
        "whatsapp_url": whatsapp_url,
        "whatsapp_client_url": whatsapp_client_url,
        "whatsapp_text": whatsapp_text,
    })


def _build_whatsapp_message(sale) -> str:
    date_fr = sale.sale_date.strftime("%d/%m/%Y")
    lines = [f"*Commande AquaTogo* — {date_fr}"]
    if sale.client:
        lines.append(f"Client : {sale.client.name}")
    lines.append("")
    for item in sale.items.all():
        if item.product:
            name = item.product.name
        elif item.service:
            name = item.service.name
        else:
            name = "Article supprimé"
        unit = f"{item.unit_price:,.0f}".replace(",", " ")
        total = f"{item.line_total:,.0f}".replace(",", " ")
        lines.append(f"- {name} x{item.quantity} ({unit} FCFA) : *{total} FCFA*")
    lines.append("")
    total_fmt = f"{sale.total_amount:,.0f}".replace(",", " ")
    lines.append(f"*Total : {total_fmt} FCFA*")
    return "\n".join(lines)


# ── sale_add_payment ──────────────────────────────────────────────────────────

@login_required
@require_POST
def sale_add_payment(request, pk):
    sale = get_object_or_404(Sale, pk=pk)
    amount = _parse_decimal(request.POST.get("amount", "0"))
    payment_method = request.POST.get("payment_method", "cash")

    if amount > Decimal("0.00"):
        remaining = sale.remaining_balance
        if amount > remaining:
            amount = remaining
        valid_methods = [m[0] for m in Payment.Method.choices]
        if payment_method not in valid_methods:
            payment_method = "cash"
        Payment.objects.create(
            sale=sale,
            recorded_by=request.user,
            amount=amount,
            payment_method=payment_method,
            payment_date=date.today(),
        )

    return redirect("sales:detail", pk=pk)


# ── sale_cancel ───────────────────────────────────────────────────────────────

@login_required
def sale_cancel(request, pk):
    sale = get_object_or_404(
        Sale.objects.select_related("client", "created_by").prefetch_related("items__product"),
        pk=pk,
    )
    if not sale.can_cancel:
        messages.error(request, "Cette vente ne peut plus être annulée.")
        return redirect("sales:detail", pk=pk)

    if request.method == "POST":
        reason = request.POST.get("reason", "").strip()
        if not reason:
            return render(request, "sales/cancel.html", {
                "sale": sale,
                "error": "La raison est obligatoire.",
            })
        try:
            sale.cancel(request.user, reason)
            messages.success(request, f"Vente #{pk} annulée. Le stock a été restauré.")
        except ValueError as e:
            messages.error(request, str(e))
        return redirect("sales:detail", pk=pk)

    return render(request, "sales/cancel.html", {"sale": sale})


# ── sale_modify ───────────────────────────────────────────────────────────────

@login_required
def sale_modify(request, pk):
    sale = get_object_or_404(
        Sale.objects.select_related("client", "created_by")
        .prefetch_related("items__product", "items__service"),
        pk=pk,
    )
    if not sale.can_modify:
        messages.error(request, "Cette vente ne peut plus être modifiée.")
        return redirect("sales:detail", pk=pk)

    clients = Client.objects.filter(is_active=True).order_by("name")
    products_qs = Product.objects.filter(is_active=True).select_related("category").order_by("name")
    services_qs = Service.objects.filter(is_active=True).order_by("name")
    products_data = [
        {"id": p.id, "name": p.name, "selling_price": str(p.selling_price),
         "purchase_price": str(p.purchase_price), "stock_quantity": p.stock_quantity}
        for p in products_qs
    ]
    services_data = [
        {"id": s.id, "name": s.name, "price": str(s.price)}
        for s in services_qs
    ]

    ctx_base = {"sale": sale, "clients": clients,
                "products_data": products_data, "services_data": services_data}

    if request.method == "POST":
        reason = request.POST.get("reason", "").strip()
        if not reason:
            return render(request, "sales/modify.html", {
                **ctx_base, "error": "La raison de modification est obligatoire.",
            })

        new_client_id = request.POST.get("client_id", "").strip()
        new_sale_date = request.POST.get("sale_date", "").strip()

        try:
            with transaction.atomic():
                # Snapshot avant modification
                snapshot = {
                    "client": sale.client.name if sale.client else None,
                    "sale_date": str(sale.sale_date),
                    "total_amount": str(sale.total_amount),
                    "items": [
                        {
                            "id": item.pk,
                            "label": str(item),
                            "quantity": item.quantity,
                            "unit_price": str(item.unit_price),
                            "line_total": str(item.line_total),
                        }
                        for item in sale.items.all()
                    ],
                }

                # Suppressions d'articles
                delete_ids = set(request.POST.getlist("delete_items"))
                for item in sale.items.select_related("product").filter(pk__in=delete_ids):
                    if item.product:
                        item.product.increase_stock(item.quantity)
                    item.delete()

                # Mise à jour quantités / prix
                for item in sale.items.select_related("product").all():
                    new_qty_raw = request.POST.get(f"item_{item.pk}_quantity", "").strip()
                    new_price_raw = request.POST.get(f"item_{item.pk}_unit_price", "").strip()
                    changed = False

                    if new_qty_raw:
                        new_qty = max(1, int(new_qty_raw))
                        if new_qty != item.quantity:
                            if item.product:
                                diff = item.quantity - new_qty
                                if diff > 0:
                                    item.product.increase_stock(diff)
                                else:
                                    item.product.decrease_stock(-diff)
                            item.quantity = new_qty
                            changed = True

                    if new_price_raw:
                        new_price = _parse_decimal(new_price_raw)
                        if new_price > Decimal("0") and new_price != item.unit_price:
                            item.unit_price = new_price
                            changed = True

                    if changed:
                        item.save()

                # Mise à jour client
                if new_client_id:
                    try:
                        sale.client = Client.objects.get(pk=int(new_client_id), is_active=True)
                    except (Client.DoesNotExist, ValueError):
                        pass
                else:
                    sale.client = None

                # Mise à jour date
                if new_sale_date:
                    try:
                        from datetime import date as date_type
                        y, m, d = new_sale_date.split("-")
                        sale.sale_date = date_type(int(y), int(m), int(d))
                    except (ValueError, AttributeError):
                        pass

                sale.save(update_fields=["client", "sale_date"])

                # Nouveaux articles
                new_count = int(request.POST.get("new_items_count", 0) or 0)
                for i in range(new_count):
                    n_type = request.POST.get(f"new_type_{i}", "").strip()
                    n_id_raw = request.POST.get(f"new_id_{i}", "").strip()
                    n_qty_raw = request.POST.get(f"new_qty_{i}", "1").strip()
                    n_price_raw = request.POST.get(f"new_price_{i}", "").strip()
                    if not n_type or not n_id_raw or not n_price_raw:
                        continue
                    n_id = int(n_id_raw)
                    n_qty = max(1, int(n_qty_raw or 1))
                    n_price = _parse_decimal(n_price_raw)
                    if n_price <= Decimal("0"):
                        continue
                    if n_type == "product":
                        prod = Product.objects.select_for_update().get(pk=n_id, is_active=True)
                        SaleItem.objects.create(
                            sale=sale, product=prod, quantity=n_qty,
                            unit_price=n_price, purchase_price_snapshot=prod.purchase_price,
                        )
                    elif n_type == "service":
                        svc = Service.objects.get(pk=n_id, is_active=True)
                        SaleItem.objects.create(
                            sale=sale, service=svc, quantity=n_qty,
                            unit_price=n_price, purchase_price_snapshot=Decimal("0.00"),
                        )

                sale.recompute_totals()
                sale.update_payment_status()

                SaleModificationLog.objects.create(
                    sale=sale,
                    modified_by=request.user,
                    reason=reason,
                    snapshot_before=snapshot,
                )

        except Exception as e:
            return render(request, "sales/modify.html", {**ctx_base, "error": str(e)})

        messages.success(request, f"Vente #{pk} modifiée avec succès.")
        return redirect("sales:detail", pk=pk)

    return render(request, "sales/modify.html", ctx_base)


# ── API views ─────────────────────────────────────────────────────────────────

@login_required
@require_POST
def api_create_client(request):
    import json as _json
    try:
        body = _json.loads(request.body)
    except (ValueError, KeyError):
        return JsonResponse({"error": "Données invalides"}, status=400)
    name = body.get("name", "").strip()
    phone = body.get("phone", "").strip()
    if not name:
        return JsonResponse({"error": "Le nom est requis"}, status=400)
    client = Client.objects.create(name=name, phone=phone)
    return JsonResponse({"id": client.id, "name": client.name, "phone": client.phone})


@login_required
@require_GET
def api_clients(request):
    q = request.GET.get("q", "").strip()
    qs = Client.objects.filter(is_active=True)
    if q:
        qs = qs.filter(name__icontains=q)
    data = list(qs.values("id", "name", "phone")[:15])
    return JsonResponse(data, safe=False)


@login_required
@require_GET
def api_products(_request):
    products = Product.objects.filter(is_active=True).select_related("category").order_by("name")
    data = [
        {
            "id": p.id,
            "name": p.name,
            "category": p.category.name,
            "selling_price": str(p.selling_price),
            "wholesale_price": str(p.wholesale_price) if p.wholesale_price is not None else None,
            "purchase_price": str(p.purchase_price),
            "stock_quantity": p.stock_quantity,
            "is_out_of_stock": p.is_out_of_stock,
        }
        for p in products
    ]
    return JsonResponse(data, safe=False)


@login_required
@require_GET
def api_services(_request):
    services = Service.objects.filter(is_active=True).order_by("name")
    data = [
        {
            "id": s.id,
            "name": s.name,
            "price": str(s.price),
        }
        for s in services
    ]
    return JsonResponse(data, safe=False)


# ── sale_export_excel ─────────────────────────────────────────────────────────

@login_required
@require_GET
def sale_export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    today = date.today()
    sales = (
        Sale.objects
        .filter(sale_date=today)
        .select_related("client", "created_by")
        .prefetch_related("items__product", "items__service")
        .order_by("pk")
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Ventes {today.strftime('%d-%m-%Y')}"

    brand_blue = "0EA5E9"
    light_blue = "E0F2FE"

    # ── Titre ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = f"AquaTogo — Résumé des ventes du {today.strftime('%d/%m/%Y')}"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=brand_blue)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── En-têtes colonnes ─────────────────────────────────────────────────────
    headers = ["#Vente", "Client", "Articles", "Statut", "Montant (FCFA)", "Bénéfice (FCFA)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="0369A1")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # ── Données ───────────────────────────────────────────────────────────────
    sales_list = list(sales)
    for i, sale in enumerate(sales_list):
        row = 3 + i
        articles = ", ".join(
            item.product.name if item.product else (item.service.name if item.service else "—")
            for item in sale.items.all()
        )
        bg = "F8FAFC" if i % 2 == 0 else "FFFFFF"
        values = [
            sale.pk,
            sale.client.name if sale.client else "Anonyme",
            articles,
            sale.get_payment_status_display(),
            float(sale.total_amount),
            float(sale.total_profit),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 3))
            if col in (5, 6):
                cell.number_format = "#,##0"
            if col == 6:
                cell.font = Font(color="15803D")

    # ── Ligne total ───────────────────────────────────────────────────────────
    total_row = 3 + len(sales_list)
    total_ca = sum(float(s.total_amount) for s in sales_list)
    total_profit = sum(float(s.total_profit) for s in sales_list)

    ws.cell(row=total_row, column=4, value="TOTAL").font = Font(bold=True, size=10)
    ws.cell(row=total_row, column=4).fill = PatternFill("solid", fgColor=light_blue)

    for col, val in [(5, total_ca), (6, total_profit)]:
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.font = Font(bold=True, size=10, color="15803D" if col == 6 else "000000")
        cell.fill = PatternFill("solid", fgColor=light_blue)
        cell.number_format = "#,##0"

    # ── Largeurs colonnes ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"ventes_{today.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ── sale_invoice_pdf ──────────────────────────────────────────────────────────

@login_required
def sale_invoice_pdf(request, pk):
    import os
    from io import BytesIO
    from django.conf import settings
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Image,
    )
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER

    sale = get_object_or_404(Sale, pk=pk)
    items = list(sale.items.select_related("product", "service").all())
    payments = list(sale.payments.all())

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        leftMargin=2.5 * cm,
        rightMargin=2.5 * cm,
    )

    brand = colors.HexColor("#0EA5E9")
    gray  = colors.HexColor("#6B7280")
    light = colors.HexColor("#F0F9FF")

    def _p(text, **kw):
        return Paragraph(text, ParagraphStyle("_", **kw))

    story = []

    # ── En-tête boutique ──────────────────────────────────────────────────────
    logo_path = os.path.join(settings.BASE_DIR, "static", "img", "logo.png")
    if os.path.exists(logo_path):
        logo_cell = Image(logo_path, width=1.4 * cm, height=1.4 * cm)
        name_cell = _p(
            '<font size="18" color="#0EA5E9"><b>AquaTogo</b></font><br/>'
            '<font size="9" color="#6B7280">Produits et services d\'aquariophilie</font>',
        )
        left_col = Table([[logo_cell, name_cell]], colWidths=[1.6 * cm, 8 * cm])
        left_col.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("LEFTPADDING", (1, 0), (1, 0), 6)]))
    else:
        left_col = _p(
            '<font size="18" color="#0EA5E9"><b>AquaTogo</b></font><br/>'
            '<font size="9" color="#6B7280">Produits et services d\'aquariophilie</font>',
        )

    header_data = [[
        left_col,
        _p(f'<font size="9" color="#6B7280">Date : {sale.sale_date.strftime("%d/%m/%Y")}<br/>'
           f'Vendeur : {sale.created_by.get_full_name() or sale.created_by.username}</font>',
           alignment=TA_RIGHT),
    ]]
    ht = Table(header_data, colWidths=[10 * cm, 7.5 * cm])
    ht.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
    story.append(ht)
    story.append(Spacer(1, 0.4 * cm))
    story.append(HRFlowable(width="100%", thickness=1.5, color=brand))
    story.append(Spacer(1, 0.5 * cm))

    # ── Numéro de facture + client ─────────────────────────────────────────────
    client_name = sale.client.name if sale.client else "Client anonyme"
    client_phone = getattr(sale.client, "phone", "") if sale.client else ""
    status_colors = {"paid": "#15803D", "partial": "#D97706", "unpaid": "#DC2626"}
    s_color = status_colors.get(sale.payment_status, "#374151")

    info_data = [[
        _p(f'<font size="14"><b>FACTURE N° {sale.pk:04d}</b></font><br/>'
           f'<font size="9" color="{s_color}"><b>● {sale.get_payment_status_display()}</b></font>'),
        _p('<b>Client</b><br/>'
           f'<font size="10">{client_name}</font>'
           + (f'<br/><font size="9" color="#6B7280">Tél : {client_phone}</font>' if client_phone else "")),
    ]]
    it = Table(info_data, colWidths=[9 * cm, 8.5 * cm])
    it.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BACKGROUND", (1, 0), (1, 0), light),
        ("LEFTPADDING", (1, 0), (1, 0), 10),
        ("TOPPADDING", (1, 0), (1, 0), 8),
        ("BOTTOMPADDING", (1, 0), (1, 0), 8),
        ("RIGHTPADDING", (1, 0), (1, 0), 10),
    ]))
    story.append(it)
    story.append(Spacer(1, 0.7 * cm))

    # ── Tableau articles ───────────────────────────────────────────────────────
    def _fmt(n):
        return f"{n:,.0f}".replace(",", "\u202f")

    rows = [[
        _p("<b>Article</b>", fontSize=9),
        _p("<b>Qté</b>", fontSize=9, alignment=TA_RIGHT),
        _p("<b>Prix unit. (FCFA)</b>", fontSize=9, alignment=TA_RIGHT),
        _p("<b>Total (FCFA)</b>", fontSize=9, alignment=TA_RIGHT),
    ]]
    for item in items:
        name = item.product.name if item.product else (item.service.name if item.service else "—")
        rows.append([
            _p(name, fontSize=9),
            _p(str(item.quantity), fontSize=9, alignment=TA_RIGHT),
            _p(_fmt(item.unit_price), fontSize=9, alignment=TA_RIGHT),
            _p(_fmt(item.line_total), fontSize=9, alignment=TA_RIGHT),
        ])
    rows.append([
        "", "",
        _p("<b>TOTAL</b>", fontSize=10, alignment=TA_RIGHT),
        _p(f"<b>{_fmt(sale.total_amount)} FCFA</b>", fontSize=10, alignment=TA_RIGHT),
    ])

    at = Table(rows, colWidths=[9.5 * cm, 2 * cm, 3.5 * cm, 3.5 * cm])
    at.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), brand),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F8FAFC")]),
        ("BACKGROUND", (0, -1), (-1, -1), light),
        ("LINEABOVE", (0, -1), (-1, -1), 1, colors.HexColor("#CBD5E1")),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(at)
    story.append(Spacer(1, 0.6 * cm))

    # ── Paiements ─────────────────────────────────────────────────────────────
    if payments:
        story.append(_p("<b>Paiements reçus</b>", fontSize=10))
        story.append(Spacer(1, 0.25 * cm))

        pay_rows = [[
            _p("<b>Date</b>", fontSize=9),
            _p("<b>Méthode</b>", fontSize=9),
            _p("<b>Montant (FCFA)</b>", fontSize=9, alignment=TA_RIGHT),
        ]]
        for p in payments:
            pay_rows.append([
                _p(p.payment_date.strftime("%d/%m/%Y"), fontSize=9),
                _p(p.get_payment_method_display(), fontSize=9),
                _p(_fmt(p.amount), fontSize=9, alignment=TA_RIGHT),
            ])
        if sale.remaining_balance > 0:
            pay_rows.append([
                "",
                _p("<b>Reste à payer</b>", fontSize=9),
                _p(f'<font color="#DC2626"><b>{_fmt(sale.remaining_balance)} FCFA</b></font>',
                   fontSize=9, alignment=TA_RIGHT),
            ])

        pt = Table(pay_rows, colWidths=[4 * cm, 7 * cm, 4 * cm])
        pt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F1F5F9")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
        ]))
        story.append(pt)

    # ── Pied de page ──────────────────────────────────────────────────────────
    story.append(Spacer(1, 1.2 * cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#E5E7EB")))
    story.append(Spacer(1, 0.3 * cm))
    story.append(_p("AquaTogo — Merci pour votre confiance !",
                    fontSize=8, textColor=gray, alignment=TA_CENTER))

    if sale.payment_status == "paid":
        def _draw_paid_stamp(canvas, doc):
            # Police 62pt : cap-height ≈ 45pt, baseline → top ≈ 45pt
            # Rect height=72 → top at 36, center at 0
            # baseline = center − cap_height/2 = 0 − 22 = −22 → top = -22+45 = 23 < 36 ✓
            canvas.saveState()
            canvas.setFillAlpha(0.30)
            canvas.setStrokeAlpha(0.30)
            stamp_green = colors.HexColor("#15803D")
            canvas.setFillColor(stamp_green)
            canvas.setStrokeColor(stamp_green)
            canvas.setLineWidth(5)
            canvas.setFont("Helvetica-Bold", 62)
            w, h = A4
            canvas.translate(w / 2, h / 2)
            canvas.rotate(42)
            # Rect centré verticalement sur 0 : de -36 à +36 (height=72)
            canvas.roundRect(-108, -36, 216, 72, 10, fill=0, stroke=1)
            # Baseline à -22 : cap top ≈ -22+45=23, bien dans le rect
            canvas.drawCentredString(0, -22, "PAYÉ")
            canvas.restoreState()
        doc.build(story, onFirstPage=_draw_paid_stamp, onLaterPages=_draw_paid_stamp)
    else:
        doc.build(story)
    buffer.seek(0)

    filename = f"Facture_AquaTogo_{sale.pk:04d}_{sale.sale_date.strftime('%Y%m%d')}.pdf"
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ── sale_export_csv ───────────────────────────────────────────────────────────

@login_required
@require_GET
def sale_export_csv(request):
    periode = request.GET.get("periode", "month")
    today = date.today()

    if periode == "week":
        start = today - timedelta(days=today.weekday())
        end = today
        label = "semaine"
    elif periode == "month":
        start = today.replace(day=1)
        end = today
        label = "mois"
    else:
        periode = "today"
        start = today
        end = today
        label = "jour"

    sales = (
        Sale.objects
        .filter(sale_date__gte=start, sale_date__lte=end)
        .select_related("client", "created_by")
        .prefetch_related("items__product", "items__service")
        .order_by("sale_date", "pk")
    )

    filename = f"ventes_{label}_{today.strftime('%Y%m%d')}.csv"
    response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response, delimiter=";")
    writer.writerow(["Date", "Client", "Articles", "Statut", "Montant (FCFA)", "Bénéfice (FCFA)"])

    for sale in sales:
        articles = ", ".join(
            item.product.name if item.product else (item.service.name if item.service else "—")
            for item in sale.items.all()
        )
        writer.writerow([
            sale.sale_date.strftime("%d/%m/%Y"),
            sale.client.name if sale.client else "Anonyme",
            articles,
            sale.get_payment_status_display(),
            str(sale.total_amount),
            str(sale.total_profit),
        ])

    return response
