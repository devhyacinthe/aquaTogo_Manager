from datetime import date, timedelta
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Case, DecimalField, ExpressionWrapper, F, Sum, Value, When
from django.db.models.functions import Coalesce
from django.shortcuts import get_object_or_404, redirect, render

from .forms import ExpenseForm
from .models import Expense


def _period_range(periode):
    today = date.today()
    if periode == "today":
        return today, today, "Aujourd'hui"
    if periode == "week":
        start = today - timedelta(days=today.weekday())
        return start, today, "Cette semaine"
    if periode == "all":
        return None, None, "Tout"
    # default: month
    start = today.replace(day=1)
    return start, today, "Ce mois"


@login_required
def expense_list(request):
    periode = request.GET.get("periode", "month")
    category_filter = request.GET.get("category", "")

    start, end, periode_label = _period_range(periode)

    qs = Expense.objects.all()
    if start:
        qs = qs.filter(expense_date__gte=start, expense_date__lte=end)
    if category_filter:
        qs = qs.filter(category=category_filter)

    total_expenses = qs.aggregate(
        total=Coalesce(Sum("amount"), Value(Decimal("0.00")), output_field=DecimalField(max_digits=14, decimal_places=2))
    )["total"]

    context = {
        "expenses": qs,
        "total_expenses": total_expenses,
        "expense_count": qs.count(),
        "periode": periode,
        "periode_label": periode_label,
        "category_filter": category_filter,
        "categories": Expense.Category.choices,
    }
    return render(request, "accounting/expense_list.html", context)


@login_required
def expense_create(request):
    if not request.user.is_superuser:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    if request.method == "POST":
        form = ExpenseForm(request.POST)
        if form.is_valid():
            expense = form.save()
            messages.success(request, f"Dépense « {expense.label} » enregistrée.")
            return redirect("accounting:list")
    else:
        form = ExpenseForm(initial={"expense_date": date.today()})

    context = {
        "form": form,
        "title": "Nouvelle dépense",
        "submit_label": "Enregistrer la dépense",
        "is_edit": False,
    }
    return render(request, "accounting/expense_form.html", context)


@login_required
def expense_edit(request, pk):
    if not request.user.is_superuser:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    expense = get_object_or_404(Expense, pk=pk)

    if request.method == "POST":
        form = ExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            form.save()
            messages.success(request, f"Dépense « {expense.label} » mise à jour.")
            return redirect("accounting:list")
    else:
        form = ExpenseForm(instance=expense)

    context = {
        "form": form,
        "expense": expense,
        "title": f"Modifier « {expense.label} »",
        "submit_label": "Mettre à jour",
        "is_edit": True,
    }
    return render(request, "accounting/expense_form.html", context)


@login_required
def expense_delete(request, pk):
    if not request.user.is_superuser:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    if request.method != "POST":
        return redirect("accounting:list")
    expense = get_object_or_404(Expense, pk=pk)
    label = expense.label
    expense.delete()
    messages.success(request, f"Dépense « {label} » supprimée.")
    return redirect("accounting:list")


@login_required
def expense_export_excel(request):
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    periode = request.GET.get("periode", "month")
    category_filter = request.GET.get("category", "")
    start, end, periode_label = _period_range(periode)

    qs = Expense.objects.all()
    if start:
        qs = qs.filter(expense_date__gte=start, expense_date__lte=end)
    if category_filter:
        qs = qs.filter(category=category_filter)
    qs = qs.order_by("expense_date", "id")

    total = qs.aggregate(
        total=Coalesce(Sum("amount"), Value(Decimal("0.00")),
                       output_field=DecimalField(max_digits=14, decimal_places=2))
    )["total"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dépenses"

    BRAND_BLUE = "1E40AF"
    LIGHT_BLUE = "DBEAFE"
    GRAY_ROW   = "F9FAFB"
    WHITE      = "FFFFFF"

    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Titre
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"Dépenses — {periode_label}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # En-têtes
    headers = ["Date", "Libellé", "Catégorie", "Note", "Montant (FCFA)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BRAND_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[2].height = 22

    category_labels = dict(Expense.Category.choices)

    for i, expense in enumerate(qs):
        row = i + 3
        fill = PatternFill("solid", fgColor=GRAY_ROW if i % 2 == 0 else WHITE)

        cells = [
            (1, expense.expense_date.strftime("%d/%m/%Y")),
            (2, expense.label),
            (3, category_labels.get(expense.category, expense.category)),
            (4, expense.note or ""),
            (5, float(expense.amount)),
        ]
        for col, val in cells:
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill
            c.font = Font(name="Calibri", size=10)
            c.border = border
            if col == 5:
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal="right")

    # Ligne total
    total_row = qs.count() + 3
    ws.merge_cells(f"A{total_row}:D{total_row}")
    label_cell = ws.cell(row=total_row, column=1, value="TOTAL")
    label_cell.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    label_cell.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    label_cell.alignment = Alignment(horizontal="right", vertical="center")
    label_cell.border = border

    total_cell = ws.cell(row=total_row, column=5, value=float(total))
    total_cell.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    total_cell.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    total_cell.number_format = '#,##0'
    total_cell.alignment = Alignment(horizontal="right")
    total_cell.border = border
    ws.row_dimensions[total_row].height = 22

    # Largeurs colonnes
    col_widths = [14, 40, 22, 35, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"depenses_{periode}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def accounting_report(request):
    import json
    import locale
    from calendar import monthrange
    from sales.models import Sale, Payment

    periode = request.GET.get("periode", "month")
    start, end, periode_label = _period_range(periode)

    _zero = Value(Decimal("0.00"))
    _df = DecimalField(max_digits=14, decimal_places=2)

    # ── Formatted month label ────────────────────────────────────────────────
    MOIS_FR = [
        "", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre",
    ]
    today = date.today()
    month_label = f"{MOIS_FR[today.month]} {today.year}"

    # CA et bénéfice basés sur les encaissements (payment_date)
    pay_qs = Payment.objects.all()
    if start:
        pay_qs = pay_qs.filter(payment_date__gte=start, payment_date__lte=end)

    # On calcule product_revenue et service_revenue par vente via Python
    # (plus simple et fiable que de sous-annoter depuis Payment)
    pay_list = list(pay_qs.select_related("sale").prefetch_related(
        "sale__items__product", "sale__items__service"
    ))

    from products.models import ProductCategory
    all_cats = list(ProductCategory.objects.order_by("name"))
    cat_revenues = {cat.slug: Decimal("0.00") for cat in all_cats}

    total_ca     = Decimal("0.00")
    gross_profit = Decimal("0.00")
    rev_products = Decimal("0.00")
    rev_services = Decimal("0.00")
    sale_ids = set()

    for payment in pay_list:
        sale = payment.sale
        total_ca += payment.amount
        sale_ids.add(sale.id)
        if sale.total_amount > 0:
            prop = payment.amount / sale.total_amount
            gross_profit += sale.total_profit * prop
            for item in sale.items.all():
                if item.product_id and item.product:
                    rev_products += prop * item.line_total
                    cat_slug = item.product.category.slug
                    if cat_slug in cat_revenues:
                        cat_revenues[cat_slug] += prop * item.line_total
                elif item.service_id:
                    rev_services += prop * item.line_total

    sale_count = len(sale_ids)
    pct_products = int(rev_products / total_ca * 100) if total_ca > 0 else 0
    pct_services = int(rev_services / total_ca * 100) if total_ca > 0 else 0

    cat_breakdown_list = [
        {
            "slug": cat.slug,
            "name": cat.name,
            "revenue": cat_revenues.get(cat.slug, Decimal("0.00")),
            "pct": round(float(cat_revenues.get(cat.slug, Decimal("0.00")) / total_ca * 100), 1) if total_ca > 0 else 0,
        }
        for cat in all_cats
    ]

    # ── Total facturé (somme des total_amount des ventes actives de la période) ──
    invoiced_qs = Sale.objects.filter(status=Sale.SaleStatus.ACTIVE)
    if start:
        invoiced_qs = invoiced_qs.filter(sale_date__gte=start, sale_date__lte=end)
    total_invoiced = invoiced_qs.aggregate(
        total=Coalesce(Sum("total_amount"), _zero, output_field=_df)
    )["total"]

    # Total number of orders (active sales in period)
    order_count = invoiced_qs.count()

    expense_qs = Expense.objects.all()
    if start:
        expense_qs = expense_qs.filter(expense_date__gte=start, expense_date__lte=end)

    operational_qs = expense_qs.exclude(category=Expense.Category.STOCK)

    total_expenses = operational_qs.aggregate(
        total=Coalesce(Sum("amount"), _zero, output_field=_df)
    )["total"]

    stock_expenses = expense_qs.filter(category=Expense.Category.STOCK).aggregate(
        total=Coalesce(Sum("amount"), _zero, output_field=_df)
    )["total"]

    # Total toutes catégories (pour l'affichage de la répartition)
    total_all_expenses = expense_qs.aggregate(
        total=Coalesce(Sum("amount"), _zero, output_field=_df)
    )["total"]

    expenses_by_category = list(
        expense_qs.values("category")
        .annotate(subtotal=Sum("amount"))
        .order_by("-subtotal")
    )

    category_labels = dict(Expense.Category.choices)
    for row in expenses_by_category:
        row["cat_label"] = category_labels.get(row["category"], row["category"])
        row["pct"] = (row["subtotal"] / total_all_expenses * 100) if total_all_expenses > 0 else Decimal("0")

    net_profit = gross_profit - stock_expenses - total_expenses
    cogs = total_ca - gross_profit
    profit_margin = (net_profit / total_ca * 100) if total_ca > 0 else Decimal("0")

    # Capital en caisse de la période = encaissements période - dépenses période
    period_cash_in_qs = Payment.objects.all()
    period_cash_out_qs = Expense.objects.all()
    if start:
        period_cash_in_qs = period_cash_in_qs.filter(payment_date__gte=start, payment_date__lte=end)
        period_cash_out_qs = period_cash_out_qs.filter(expense_date__gte=start, expense_date__lte=end)

    cash_in_period = period_cash_in_qs.aggregate(
        total=Coalesce(Sum("amount"), _zero, output_field=_df)
    )["total"]
    cash_out_period = period_cash_out_qs.aggregate(
        total=Coalesce(Sum("amount"), _zero, output_field=_df)
    )["total"]
    capital_periode = cash_in_period - cash_out_period

    # Remaining to collect = total invoiced - total collected
    remaining_to_collect = total_invoiced - cash_in_period

    # Capital en caisse total (cumul tous temps)
    all_time_cash_in = Payment.objects.aggregate(
        total=Coalesce(Sum("amount"), _zero, output_field=_df)
    )["total"]
    all_time_cash_out = Expense.objects.aggregate(
        total=Coalesce(Sum("amount"), _zero, output_field=_df)
    )["total"]
    capital_actuel = all_time_cash_in - all_time_cash_out

    # ── Weekly evolution (revenue per week within period) ─────────────────────
    weekly_labels = []
    weekly_values = []
    if start and end:
        # Compute week boundaries within the period
        week_num = 1
        week_start = start
        while week_start <= end:
            # End of the week = next Sunday or end of period, whichever comes first
            days_to_sunday = 6 - week_start.weekday()
            week_end = min(week_start + timedelta(days=days_to_sunday), end)
            # Sum payments in this week
            week_total = Payment.objects.filter(
                payment_date__gte=week_start,
                payment_date__lte=week_end,
            ).aggregate(
                total=Coalesce(Sum("amount"), _zero, output_field=_df)
            )["total"]
            weekly_labels.append(f"Sem. {week_num}")
            weekly_values.append(int(week_total))
            week_start = week_end + timedelta(days=1)
            week_num += 1
    else:
        # For "all" period, group by month for recent 6 months
        for i in range(5, -1, -1):
            m_date = today.replace(day=1) - timedelta(days=i * 30)
            m_start = m_date.replace(day=1)
            _, last_day = monthrange(m_start.year, m_start.month)
            m_end = m_start.replace(day=last_day)
            m_total = Payment.objects.filter(
                payment_date__gte=m_start,
                payment_date__lte=m_end,
            ).aggregate(
                total=Coalesce(Sum("amount"), _zero, output_field=_df)
            )["total"]
            weekly_labels.append(MOIS_FR[m_start.month][:3])
            weekly_values.append(int(m_total))

    # Find max for chart scaling
    weekly_max = max(weekly_values) if weekly_values else 1

    # ── KPIs avancés ──────────────────────────────────────────────────────────

    # Nombre de jours dans la période
    if start and end:
        nb_days = max((end - start).days + 1, 1)
    else:
        # "Tout" : du premier paiement à aujourd'hui
        first_payment = Payment.objects.order_by("payment_date").first()
        if first_payment:
            nb_days = max((today - first_payment.payment_date).days + 1, 1)
        else:
            nb_days = 1

    avg_daily_revenue = total_ca / nb_days if nb_days > 0 else Decimal("0")
    avg_ticket = total_ca / order_count if order_count > 0 else Decimal("0")
    collection_rate = (cash_in_period / total_invoiced * 100) if total_invoiced > 0 else Decimal("0")

    # Clients uniques sur la période
    unique_clients = invoiced_qs.values("client").distinct().count()

    # ── CA mensuel & annuel (référence) ───────────────────────────────────
    month_start = today.replace(day=1)
    _, last_day_of_month = monthrange(today.year, today.month)
    month_end = today.replace(day=last_day_of_month)
    ca_mensuel = Sale.objects.filter(
        status=Sale.SaleStatus.ACTIVE,
        sale_date__gte=month_start,
        sale_date__lte=month_end,
    ).aggregate(total=Coalesce(Sum("total_amount"), _zero, output_field=_df))["total"]

    year_start = today.replace(month=1, day=1)
    ca_annuel = Sale.objects.filter(
        status=Sale.SaleStatus.ACTIVE,
        sale_date__gte=year_start,
        sale_date__lte=today,
    ).aggregate(total=Coalesce(Sum("total_amount"), _zero, output_field=_df))["total"]

    # ── Tableau détaillé par semaine ──────────────────────────────────────
    weekly_detail = []
    if start and end:
        week_num = 1
        week_start = start
        while week_start <= end:
            days_to_sunday = 6 - week_start.weekday()
            week_end = min(week_start + timedelta(days=days_to_sunday), end)

            w_payments = Payment.objects.filter(
                payment_date__gte=week_start,
                payment_date__lte=week_end,
            )
            w_ca = w_payments.aggregate(
                total=Coalesce(Sum("amount"), _zero, output_field=_df)
            )["total"]
            w_orders = Sale.objects.filter(
                status=Sale.SaleStatus.ACTIVE,
                sale_date__gte=week_start,
                sale_date__lte=week_end,
            ).count()
            w_expenses = Expense.objects.filter(
                expense_date__gte=week_start,
                expense_date__lte=week_end,
            ).aggregate(
                total=Coalesce(Sum("amount"), _zero, output_field=_df)
            )["total"]

            weekly_detail.append({
                "label": f"Sem. {week_num}",
                "dates": f"{week_start.strftime('%d/%m')} – {week_end.strftime('%d/%m')}",
                "ca": w_ca,
                "orders": w_orders,
                "avg_ticket": (w_ca / w_orders) if w_orders > 0 else Decimal("0"),
                "expenses": w_expenses,
                "solde": w_ca - w_expenses,
                "solde_positif": (w_ca - w_expenses) >= 0,
            })
            week_start = week_end + timedelta(days=1)
            week_num += 1

    # ── Comparaison avec la période précédente ────────────────────────────
    prev_ca = Decimal("0")
    ca_variation = None
    if start and end:
        delta = (end - start).days + 1
        prev_start = start - timedelta(days=delta)
        prev_end = start - timedelta(days=1)
        prev_ca = Payment.objects.filter(
            payment_date__gte=prev_start,
            payment_date__lte=prev_end,
        ).aggregate(
            total=Coalesce(Sum("amount"), _zero, output_field=_df)
        )["total"]
        if prev_ca > 0:
            ca_variation = round(float((cash_in_period - prev_ca) / prev_ca * 100), 1)

    context = {
        "total_ca": total_ca,
        "gross_profit": gross_profit,
        "cogs": cogs,
        "stock_expenses": stock_expenses,
        "total_expenses": total_expenses,
        "net_profit": net_profit,
        "profit_margin": profit_margin,
        "is_profitable": net_profit >= 0,
        "expenses_by_category": expenses_by_category,
        "sale_count": sale_count,
        "order_count": order_count,
        "expense_count": expense_qs.count(),
        "periode": periode,
        "periode_label": periode_label,
        "month_label": month_label,
        "rev_products":      rev_products,
        "rev_services":      rev_services,
        "pct_products":      pct_products,
        "pct_services":      pct_services,
        "cat_breakdown_list": cat_breakdown_list,
        "total_invoiced": total_invoiced,
        "cash_in_period": cash_in_period,
        "cash_out_period": cash_out_period,
        "remaining_to_collect": remaining_to_collect,
        "capital_periode": capital_periode,
        "capital_periode_positif": capital_periode >= 0,
        "capital_actuel": capital_actuel,
        "capital_positif": capital_actuel >= 0,
        # Chart data
        "weekly_labels_json": json.dumps(weekly_labels),
        "weekly_values_json": json.dumps(weekly_values),
        "weekly_max": weekly_max,
        # KPIs avancés
        "nb_days": nb_days,
        "avg_daily_revenue": avg_daily_revenue,
        "avg_ticket": avg_ticket,
        "unique_clients": unique_clients,
        "collection_rate": collection_rate,
        "ca_mensuel": ca_mensuel,
        "ca_annuel": ca_annuel,
        # Tableau détaillé
        "weekly_detail": weekly_detail,
        # Comparaison
        "ca_variation": ca_variation,
        "prev_ca": prev_ca,
        "is_superuser": request.user.is_superuser,
    }
    return render(request, "accounting/report.html", context)


@login_required
def report_pdf(request):
    import os
    from io import BytesIO
    from django.conf import settings
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        HRFlowable, Image, KeepTogether,
    )
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT

    from sales.models import Payment

    periode = request.GET.get("periode", "month")
    start, end, periode_label = _period_range(periode)

    _zero = Value(Decimal("0.00"))
    _df = DecimalField(max_digits=14, decimal_places=2)

    pay_qs = Payment.objects.all()
    if start:
        pay_qs = pay_qs.filter(payment_date__gte=start, payment_date__lte=end)

    pay_list = list(pay_qs.select_related("sale").prefetch_related(
        "sale__items__product", "sale__items__service"
    ))

    from products.models import ProductCategory
    all_cats_pdf = list(ProductCategory.objects.order_by("name"))
    cat_revenues_pdf = {cat.slug: Decimal("0.00") for cat in all_cats_pdf}

    total_ca     = Decimal("0.00")
    gross_profit = Decimal("0.00")
    rev_products  = Decimal("0.00")
    rev_services  = Decimal("0.00")
    sale_ids = set()

    for payment in pay_list:
        sale = payment.sale
        total_ca += payment.amount
        sale_ids.add(sale.id)
        if sale.total_amount > 0:
            prop = payment.amount / sale.total_amount
            gross_profit += sale.total_profit * prop
            for item in sale.items.all():
                if item.product_id and item.product:
                    rev_products += prop * item.line_total
                    cat_slug = item.product.category.slug
                    if cat_slug in cat_revenues_pdf:
                        cat_revenues_pdf[cat_slug] += prop * item.line_total
                elif item.service_id:
                    rev_services += prop * item.line_total

    cat_breakdown_pdf = [
        {
            "slug": cat.slug,
            "name": cat.name,
            "revenue": cat_revenues_pdf.get(cat.slug, Decimal("0.00")),
        }
        for cat in all_cats_pdf
    ]

    expense_qs = Expense.objects.all()
    if start:
        expense_qs = expense_qs.filter(expense_date__gte=start, expense_date__lte=end)

    operational_qs_pdf = expense_qs.exclude(category=Expense.Category.STOCK)

    total_expenses = operational_qs_pdf.aggregate(
        total=Coalesce(Sum("amount"), _zero, output_field=_df)
    )["total"]

    stock_expenses = expense_qs.filter(category=Expense.Category.STOCK).aggregate(
        total=Coalesce(Sum("amount"), _zero, output_field=_df)
    )["total"]

    total_all_expenses_pdf = expense_qs.aggregate(
        total=Coalesce(Sum("amount"), _zero, output_field=_df)
    )["total"]

    expenses_by_category = list(
        expense_qs.values("category")
        .annotate(subtotal=Sum("amount"))
        .order_by("-subtotal")
    )
    category_labels = dict(Expense.Category.choices)
    for row in expenses_by_category:
        row["cat_label"] = category_labels.get(row["category"], row["category"])

    net_profit    = gross_profit - stock_expenses - total_expenses
    cogs          = total_ca - gross_profit
    profit_margin = (net_profit / total_ca * 100) if total_ca > 0 else Decimal("0")

    # ── ReportLab ─────────────────────────────────────────────────────────────
    brand  = colors.HexColor("#0EA5E9")
    green  = colors.HexColor("#16A34A")
    red    = colors.HexColor("#DC2626")
    gray   = colors.HexColor("#6B7280")
    light  = colors.HexColor("#F0F9FF")
    white  = colors.white

    def _p(text, **kw):
        return Paragraph(text, ParagraphStyle("_", **kw))

    def _fmt(n):
        return f"{int(n):,}".replace(",", "\u202f")

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=2 * cm, bottomMargin=2 * cm,
        leftMargin=2 * cm, rightMargin=2 * cm,
    )

    story = []

    # ── En-tête ───────────────────────────────────────────────────────────────
    logo_path = os.path.join(settings.BASE_DIR, "static", "img", "logo.png")
    if os.path.exists(logo_path):
        logo_cell = Image(logo_path, width=1.4 * cm, height=1.4 * cm)
        name_cell = _p(
            '<font size="16" color="#0EA5E9"><b>AquaTogo</b></font><br/>'
            '<font size="8" color="#6B7280">Produits et services d\'aquariophilie</font>',
        )
        left_col = Table([[logo_cell, name_cell]], colWidths=[1.6 * cm, 8 * cm])
        left_col.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (1, 0), (1, 0), 6),
        ]))
    else:
        left_col = _p(
            '<font size="16" color="#0EA5E9"><b>AquaTogo</b></font><br/>'
            '<font size="8" color="#6B7280">Produits et services d\'aquariophilie</font>',
        )

    from datetime import date as _date
    today_str = _date.today().strftime("%d/%m/%Y")
    header_table = Table(
        [[left_col, _p(
            f'<font size="9" color="#6B7280">Rapport généré le {today_str}<br/>'
            f'Période : {periode_label}</font>',
            alignment=TA_RIGHT,
        )]],
        colWidths=[10 * cm, 7 * cm],
    )
    header_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
    story.append(header_table)
    story.append(Spacer(1, 0.3 * cm))
    story.append(HRFlowable(width="100%", thickness=2, color=brand))
    story.append(Spacer(1, 0.4 * cm))

    # Titre rapport
    story.append(_p(
        f'<font size="14"><b>Rapport Comptable</b></font>   '
        f'<font size="10" color="#6B7280">{periode_label}</font>',
    ))
    story.append(Spacer(1, 0.5 * cm))

    # ── KPI bandeaux ─────────────────────────────────────────────────────────
    net_color = "#16A34A" if net_profit >= 0 else "#DC2626"
    net_sign  = "+" if net_profit >= 0 else ""
    kpi_data = [[
        _p(f'<font size="8" color="#FFFFFF">Chiffre d\'affaires</font><br/>'
           f'<font size="16" color="#FFFFFF"><b>{_fmt(total_ca)}</b></font><br/>'
           f'<font size="7" color="#BAE6FD">FCFA · {len(sale_ids)} vente{"s" if len(sale_ids)>1 else ""}</font>',
           alignment=TA_CENTER),
        _p(f'<font size="8" color="#374151">Bénéfice brut</font><br/>'
           f'<font size="16" color="#374151"><b>{_fmt(gross_profit)}</b></font><br/>'
           f'<font size="7" color="#6B7280">FCFA · après coût marchand.</font>',
           alignment=TA_CENTER),
        _p(f'<font size="8" color="{net_color}">Résultat net</font><br/>'
           f'<font size="16" color="{net_color}"><b>{net_sign}{_fmt(net_profit)}</b></font><br/>'
           f'<font size="7" color="{net_color}">{float(profit_margin):.1f}% marge nette</font>',
           alignment=TA_CENTER),
    ]]
    kpi_t = Table(kpi_data, colWidths=[5.5 * cm, 5.5 * cm, 5.5 * cm])
    kpi_t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), brand),
        ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#F8FAFC")),
        ("BACKGROUND", (2, 0), (2, 0), colors.HexColor("#F0FDF4") if net_profit >= 0 else colors.HexColor("#FEF2F2")),
        ("ROUNDEDCORNERS", [4]),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        ("BOX", (0, 0), (0, 0), 0, colors.transparent),
        ("BOX", (1, 0), (1, 0), 0.5, colors.HexColor("#E2E8F0")),
        ("BOX", (2, 0), (2, 0), 0.5, colors.HexColor("#E2E8F0")),
        ("LEFTPADDING", (1, 0), (1, 0), 10),
        ("LEFTPADDING", (2, 0), (2, 0), 10),
    ]))
    story.append(kpi_t)
    story.append(Spacer(1, 0.5 * cm))

    # ── Compte de résultat ────────────────────────────────────────────────────
    story.append(_p("<b>Compte de résultat</b>", fontSize=10))
    story.append(Spacer(1, 0.2 * cm))

    pl_rows = [
        ["Chiffre d'affaires",          f"{_fmt(total_ca)} FCFA",             "", False],
        ["Coût des articles vendus",    f"− {_fmt(cogs)} FCFA",               "", False],
        ["Bénéfice brut",               f"{_fmt(gross_profit)} FCFA",         "", True],
        ["Achats de stock",             f"− {_fmt(stock_expenses)} FCFA",     "", False],
        ["Dépenses opérationnelles",    f"− {_fmt(total_expenses)} FCFA",     "", False],
        ["RÉSULTAT NET",                f"{net_sign}{_fmt(net_profit)} FCFA", f"{float(profit_margin):.1f}%", True],
    ]
    net_row_idx = len(pl_rows) - 1
    gross_row_idx = 2

    def _pl_cell(text, bold=False, color="#374151", align=TA_LEFT):
        return _p(f'<font color="{color}">{"<b>" if bold else ""}{text}{"</b>" if bold else ""}</font>',
                  fontSize=9, alignment=align)

    pl_table_data = [
        [_pl_cell(r[0], bold=r[3]),
         _pl_cell(r[1], bold=r[3],
           color=("#DC2626" if "−" in r[1] else ("#16A34A" if i == net_row_idx and net_profit >= 0 else "#DC2626" if i == net_row_idx else "#374151")),
           align=TA_RIGHT),
         _pl_cell(r[2], color="#6B7280", align=TA_RIGHT)]
        for i, r in enumerate(pl_rows)
    ]

    pl_t = Table(pl_table_data, colWidths=[9 * cm, 5 * cm, 2.5 * cm])
    pl_style = [
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        ("LINEBELOW", (0, 0), (-1, -2), 0.5, colors.HexColor("#E5E7EB")),
        ("BACKGROUND", (0, gross_row_idx), (-1, gross_row_idx), colors.HexColor("#F8FAFC")),
        ("BACKGROUND", (0, net_row_idx), (-1, net_row_idx),
         colors.HexColor("#F0FDF4") if net_profit >= 0 else colors.HexColor("#FEF2F2")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
    ]
    pl_t.setStyle(TableStyle(pl_style))
    story.append(pl_t)
    story.append(Spacer(1, 0.5 * cm))

    # ── Répartition des revenus ───────────────────────────────────────────────
    if total_ca > 0:
        story.append(_p("<b>Répartition des revenus</b>", fontSize=10))
        story.append(Spacer(1, 0.2 * cm))

        rev_rows = []
        _cat_hex = {"fish": "#0EA5E9", "accessory": "#8B5CF6", "aquarium": "#F59E0B"}
        cat_items = [
            (c["name"], c["revenue"], _cat_hex.get(c["slug"], "#6B7280"))
            for c in cat_breakdown_pdf
        ] + [("Prestations", rev_services, "#10B981")]
        for label, amount, color in cat_items:
            if amount > 0:
                pct = float(amount / total_ca * 100)
                rev_rows.append([
                    _p(f'<font color="{color}">■</font>  <font size="9">{label}</font>', fontSize=9),
                    _p(f'<font size="9">{_fmt(amount)} FCFA</font>', fontSize=9, alignment=TA_RIGHT),
                    _p(f'<font size="9" color="#6B7280">{pct:.0f}%</font>', fontSize=9, alignment=TA_RIGHT),
                ])

        if rev_rows:
            rev_t = Table(rev_rows, colWidths=[8 * cm, 5.5 * cm, 3 * cm])
            rev_t.setStyle(TableStyle([
                ("TOPPADDING",    (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING",   (0, 0), (-1, -1), 8),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
                ("LINEBELOW", (0, 0), (-1, -2), 0.5, colors.HexColor("#F1F5F9")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ]))
            story.append(rev_t)
        story.append(Spacer(1, 0.5 * cm))

    # ── Dépenses par catégorie ────────────────────────────────────────────────
    if expenses_by_category:
        story.append(_p("<b>Dépenses par catégorie</b>", fontSize=10))
        story.append(Spacer(1, 0.2 * cm))

        exp_rows = [[
            _p("<b>Catégorie</b>", fontSize=9),
            _p("<b>Montant</b>", fontSize=9, alignment=TA_RIGHT),
            _p("<b>%</b>", fontSize=9, alignment=TA_RIGHT),
        ]]
        for row in expenses_by_category:
            pct = float(row["subtotal"] / total_expenses * 100) if total_expenses > 0 else 0
            exp_rows.append([
                _p(row["cat_label"], fontSize=9),
                _p(f'{_fmt(row["subtotal"])} FCFA', fontSize=9, alignment=TA_RIGHT),
                _p(f'{pct:.0f}%', fontSize=9, alignment=TA_RIGHT),
            ])
        exp_rows.append([
            _p("<b>Total</b>", fontSize=9),
            _p(f'<b>{_fmt(total_expenses)} FCFA</b>', fontSize=9, alignment=TA_RIGHT),
            _p("", fontSize=9),
        ])

        exp_t = Table(exp_rows, colWidths=[8 * cm, 5.5 * cm, 3 * cm])
        exp_t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F1F5F9")),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F8FAFC")),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
            ("LINEBELOW", (0, 0), (-1, -2), 0.5, colors.HexColor("#E5E7EB")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#FAFAFA")]),
        ]))
        story.append(exp_t)
        story.append(Spacer(1, 0.5 * cm))

    # ── Pied de page ─────────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#E5E7EB")))
    story.append(Spacer(1, 0.25 * cm))
    story.append(_p(
        f'AquaTogo — Rapport généré le {today_str}',
        fontSize=7, textColor=gray, alignment=TA_CENTER,
    ))

    doc.build(story)
    buffer.seek(0)

    filename = f"rapport_comptable_{periode}_{_date.today().strftime('%Y%m%d')}.pdf"
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
